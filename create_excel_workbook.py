#!/usr/bin/env python3
"""
Beauty Pro Inventory System - Excel Workbook Creator
Creates a comprehensive Excel workbook with all 9 worksheets, professional formatting, and sample data.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import csv
from datetime import datetime, timedelta

# Define color scheme based on the Beauty Pro brand
class BeautyProColors:
    DEEP_CHARCOAL = "2C3E50"
    ROSE_GOLD = "E8B4B8"
    CREAM_WHITE = "F8F9FA"
    SUCCESS = "27AE60"
    WARNING = "F39C12"
    CRITICAL = "E74C3C"
    PROGRESS = "3498DB"
    WARM_GRAY = "95A5A6"
    PURE_WHITE = "FFFFFF"
    LIGHT_BORDER = "DEE2E6"
    DARK_TEXT = "566573"

def create_excel_workbook():
    """Create the complete Beauty Pro Inventory System Excel workbook"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define sheet order and data
    sheets_data = [
        ("Dashboard", "Dashboard.csv"),
        ("Categories", "Categories.csv"),
        ("Suppliers", "Suppliers.csv"),
        ("Products", "Products.csv"),
        ("Inventory", "Inventory.csv"),
        ("QuickAdd", "QuickAdd.csv"),
        ("Reorder", "Reorder.csv"),
        ("Analytics", "Analytics.csv"),
        ("Instructions", "Instructions.csv")
    ]
    
    base_path = "/home/grig/Projects/inventory_template/sheets/"
    
    for sheet_name, csv_file in sheets_data:
        print(f"Creating {sheet_name} worksheet...")
        
        # Create worksheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Read CSV data
        csv_path = os.path.join(base_path, csv_file)
        data = read_csv_data(csv_path)
        
        # Add data to worksheet
        populate_worksheet(ws, data, sheet_name)
        
        # Apply formatting
        format_worksheet(ws, sheet_name)
    
    # Add formulas and validation after all sheets are created
    add_formulas_and_validation(wb)
    
    # Save workbook
    output_path = "/home/grig/Projects/inventory_template/Beauty_Pro_Inventory_System.xlsx"
    wb.save(output_path)
    print(f"Excel workbook saved to: {output_path}")
    
    return output_path

def read_csv_data(csv_path):
    """Read CSV data and return as list of lists"""
    data = []
    try:
        with open(csv_path, 'r', encoding='utf-8') as file:
            # Read the raw CSV to preserve formatting
            reader = csv.reader(file)
            for row in reader:
                data.append(row)
    except Exception as e:
        print(f"Error reading {csv_path}: {e}")
        return []
    
    return data

def populate_worksheet(ws, data, sheet_name):
    """Populate worksheet with data"""
    if not data:
        return
    
    # Add data row by row
    for row_idx, row_data in enumerate(data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            if cell_value:  # Only add non-empty values
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

def format_worksheet(ws, sheet_name):
    """Apply professional formatting to worksheet"""
    
    # Define fonts
    header_font = Font(name='Arial', size=14, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    subheader_font = Font(name='Arial', size=12, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    body_font = Font(name='Arial', size=10, color=BeautyProColors.DARK_TEXT)
    
    # Define fills
    header_fill = PatternFill(start_color=BeautyProColors.ROSE_GOLD, end_color=BeautyProColors.ROSE_GOLD, fill_type="solid")
    subheader_fill = PatternFill(start_color=BeautyProColors.CREAM_WHITE, end_color=BeautyProColors.CREAM_WHITE, fill_type="solid")
    success_fill = PatternFill(start_color=BeautyProColors.SUCCESS, end_color=BeautyProColors.SUCCESS, fill_type="solid")
    warning_fill = PatternFill(start_color=BeautyProColors.WARNING, end_color=BeautyProColors.WARNING, fill_type="solid")
    critical_fill = PatternFill(start_color=BeautyProColors.CRITICAL, end_color=BeautyProColors.CRITICAL, fill_type="solid")
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin', color=BeautyProColors.LIGHT_BORDER),
        right=Side(style='thin', color=BeautyProColors.LIGHT_BORDER),
        top=Side(style='thin', color=BeautyProColors.LIGHT_BORDER),
        bottom=Side(style='thin', color=BeautyProColors.LIGHT_BORDER)
    )
    
    # Apply sheet-specific formatting
    if sheet_name == "Dashboard":
        format_dashboard(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, success_fill, warning_fill, critical_fill, thin_border)
    elif sheet_name == "Categories":
        format_categories(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border)
    elif sheet_name == "Suppliers":
        format_suppliers(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border)
    elif sheet_name == "Products":
        format_products(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border)
    elif sheet_name == "Inventory":
        format_inventory(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, success_fill, warning_fill, critical_fill, thin_border)
    elif sheet_name == "QuickAdd":
        format_quickadd(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border)
    elif sheet_name == "Reorder":
        format_reorder(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, success_fill, warning_fill, critical_fill, thin_border)
    elif sheet_name == "Analytics":
        format_analytics(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, success_fill, warning_fill, critical_fill, thin_border)
    elif sheet_name == "Instructions":
        format_instructions(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border)
    
    # Auto-adjust column widths
    adjust_column_widths(ws)

def format_dashboard(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, success_fill, warning_fill, critical_fill, thin_border):
    """Format Dashboard worksheet"""
    
    # Main header
    ws['A2'].font = Font(name='Arial', size=16, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    ws['A2'].fill = header_fill
    ws.merge_cells('A2:J2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Section headers
    section_headers = ['A4', 'F4', 'I4', 'A11', 'D11', 'G11', 'A20', 'D20', 'G20', 'A27', 'D27', 'G27']
    for cell in section_headers:
        if ws[cell].value:
            ws[cell].font = subheader_font
            ws[cell].fill = subheader_fill
    
    # Status indicators - color code based on content
    for row in ws.iter_rows(min_row=6, max_row=32):
        for cell in row:
            if cell.value:
                if '🟢' in str(cell.value) or 'Healthy' in str(cell.value):
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif '🟡' in str(cell.value) or 'Low Stock' in str(cell.value):
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                elif '🔴' in str(cell.value) or 'Out of Stock' in str(cell.value):
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

def format_categories(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border):
    """Format Categories worksheet"""
    
    # Main header
    ws['A2'].font = Font(name='Arial', size=16, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    ws['A2'].fill = header_fill
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data table header
    for col in range(1, 9):  # A to H
        cell = ws.cell(row=4, column=col)
        if cell.value:
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
    
    # Data rows
    for row in range(6, 12):  # Data rows
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.font = body_font
                cell.border = thin_border

def format_suppliers(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border):
    """Format Suppliers worksheet"""
    
    # Main header
    ws['A2'].font = Font(name='Arial', size=16, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    ws['A2'].fill = header_fill
    ws.merge_cells('A2:J2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data table header
    for col in range(1, 11):  # A to J
        cell = ws.cell(row=4, column=col)
        if cell.value:
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
    
    # Data rows
    for row in range(6, 10):  # Data rows
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.font = body_font
                cell.border = thin_border

def format_products(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border):
    """Format Products worksheet"""
    
    # Main header
    ws['A2'].font = Font(name='Arial', size=16, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    ws['A2'].fill = header_fill
    ws.merge_cells('A2:N2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data table header
    for col in range(1, 15):  # A to N
        cell = ws.cell(row=4, column=col)
        if cell.value:
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
    
    # Data rows
    for row in range(6, 16):  # Data rows
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.font = body_font
                cell.border = thin_border

def format_inventory(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, success_fill, warning_fill, critical_fill, thin_border):
    """Format Inventory worksheet"""
    
    # Main header
    ws['A2'].font = Font(name='Arial', size=16, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    ws['A2'].fill = header_fill
    ws.merge_cells('A2:O2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data table header
    for col in range(1, 16):  # A to O
        cell = ws.cell(row=4, column=col)
        if cell.value:
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
    
    # Data rows with status coloring
    for row in range(6, 16):  # Data rows
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.font = body_font
                cell.border = thin_border
                
                # Color code status column
                if col == 6:  # Status column
                    if '🟢' in str(cell.value) or 'Healthy' in str(cell.value):
                        cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                    elif '🟡' in str(cell.value) or 'Low Stock' in str(cell.value) or 'At Minimum' in str(cell.value):
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    elif '🔴' in str(cell.value) or 'Out of Stock' in str(cell.value):
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

def format_quickadd(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border):
    """Format QuickAdd worksheet"""
    
    # Main header
    ws['A2'].font = Font(name='Arial', size=16, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    ws['A2'].fill = header_fill
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Section headers
    section_headers = ['A4', 'A11', 'A24']
    for cell_addr in section_headers:
        cell = ws[cell_addr]
        if cell.value:
            cell.font = subheader_font
            cell.fill = subheader_fill

def format_reorder(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, success_fill, warning_fill, critical_fill, thin_border):
    """Format Reorder worksheet"""
    
    # Main header
    ws['A2'].font = Font(name='Arial', size=16, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    ws['A2'].fill = header_fill
    ws.merge_cells('A2:J2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Data table headers
    for col in range(1, 11):  # A to J
        cell = ws.cell(row=6, column=col)
        if cell.value:
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
    
    # Priority coloring in reorder table
    for row in range(8, 12):  # Data rows
        priority_cell = ws.cell(row=row, column=9)  # Priority column
        if priority_cell.value:
            if '🔴' in str(priority_cell.value) or 'URGENT' in str(priority_cell.value):
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            elif '🟡' in str(priority_cell.value) or 'MEDIUM' in str(priority_cell.value):
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

def format_analytics(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, success_fill, warning_fill, critical_fill, thin_border):
    """Format Analytics worksheet"""
    
    # Main header
    ws['A2'].font = Font(name='Arial', size=16, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    ws['A2'].fill = header_fill
    ws.merge_cells('A2:J2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Section headers
    section_headers = ['A4', 'A12', 'A22', 'A31', 'A38', 'A46', 'A54', 'A64', 'A72']
    for cell_addr in section_headers:
        cell = ws[cell_addr]
        if cell.value:
            cell.font = subheader_font
            cell.fill = subheader_fill

def format_instructions(ws, header_font, subheader_font, body_font, header_fill, subheader_fill, thin_border):
    """Format Instructions worksheet"""
    
    # Main header
    ws['A2'].font = Font(name='Arial', size=16, bold=True, color=BeautyProColors.DEEP_CHARCOAL)
    ws['A2'].fill = header_fill
    ws.merge_cells('A2:J2')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Section headers
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and ('🎯' in str(cell.value) or '📱' in str(cell.value) or '🔧' in str(cell.value) or '🎨' in str(cell.value) or '🆘' in str(cell.value) or '📞' in str(cell.value)):
                cell.font = subheader_font
                cell.fill = subheader_fill

def adjust_column_widths(ws):
    """Auto-adjust column widths based on content"""
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

def add_formulas_and_validation(wb):
    """Add Excel formulas and basic validation"""
    
    # Add some basic formulas to Products sheet for margin calculation
    products_ws = wb["Products"]
    
    # Add margin calculation formula to existing products
    for row in range(6, 16):  # Product data rows
        cost_col = 6  # F column (Cost)
        price_col = 7  # G column (Retail Price) 
        margin_col = 8  # H column (Margin %)
        
        cost_cell = get_column_letter(cost_col) + str(row)
        price_cell = get_column_letter(price_col) + str(row)
        margin_cell = get_column_letter(margin_col) + str(row)
        
        # Only add formula if both cost and price exist
        if products_ws[cost_cell].value and products_ws[price_cell].value:
            try:
                cost_val = float(str(products_ws[cost_cell].value).replace('$', ''))
                price_val = float(str(products_ws[price_cell].value).replace('$', ''))
                if price_val > 0:
                    margin = ((price_val - cost_val) / price_val) * 100
                    products_ws[margin_cell].value = f"{margin:.1f}%"
            except:
                pass

if __name__ == "__main__":
    print("Creating Beauty Pro Inventory System Excel Workbook...")
    output_file = create_excel_workbook()
    print(f"Workbook created successfully: {output_file}")