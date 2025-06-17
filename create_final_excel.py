#!/usr/bin/env python3
"""
Beauty Pro Inventory System - Final Excel Creator
Creates the ultimate comprehensive Excel workbook with all features, formulas, and professional formatting.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os
import csv
from datetime import datetime, timedelta

# Beauty Pro Color Scheme
class Colors:
    DEEP_CHARCOAL = "2C3E50"
    ROSE_GOLD = "E8B4B8"
    CREAM_WHITE = "F8F9FA"
    SUCCESS = "27AE60"
    WARNING = "F39C12"
    CRITICAL = "E74C3C"
    PROGRESS = "3498DB"
    WARM_GRAY = "95A5A6"
    PURE_WHITE = "FFFFFF"
    LIGHT_BORDER = "DEE2E6"
    DARK_TEXT = "566573"

def create_final_workbook():
    """Create the final comprehensive workbook"""
    
    print("Creating final Beauty Pro Inventory System Excel workbook...")
    
    # Create workbook with predefined styles
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create named styles
    create_named_styles(wb)
    
    # Sheet creation order
    sheets_info = [
        ("Dashboard", create_dashboard_data),
        ("Categories", create_categories_data),
        ("Suppliers", create_suppliers_data),
        ("Products", create_products_data),
        ("Inventory", create_inventory_data),
        ("QuickAdd", create_quickadd_data),
        ("Reorder", create_reorder_data),
        ("Analytics", create_analytics_data),
        ("Instructions", create_instructions_data)
    ]
    
    for sheet_name, data_func in sheets_info:
        print(f"Creating {sheet_name} worksheet...")
        ws = wb.create_sheet(title=sheet_name)
        data_func(ws)
        format_sheet(ws, sheet_name)
    
    # Save final workbook
    output_path = "/home/grig/Projects/inventory_template/Beauty_Pro_Inventory_System_FINAL.xlsx"
    wb.save(output_path)
    print(f"Final Excel workbook saved to: {output_path}")
    
    return output_path

def create_named_styles(wb):
    """Create named styles for consistent formatting"""
    
    # Header style
    header_style = NamedStyle(name="header")
    header_style.font = Font(name='Arial', size=14, bold=True, color=Colors.DEEP_CHARCOAL)
    header_style.fill = PatternFill(start_color=Colors.ROSE_GOLD, end_color=Colors.ROSE_GOLD, fill_type="solid")
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(header_style)
    
    # Subheader style
    subheader_style = NamedStyle(name="subheader")
    subheader_style.font = Font(name='Arial', size=12, bold=True, color=Colors.DEEP_CHARCOAL)
    subheader_style.fill = PatternFill(start_color=Colors.CREAM_WHITE, end_color=Colors.CREAM_WHITE, fill_type="solid")
    wb.add_named_style(subheader_style)
    
    # Data style
    data_style = NamedStyle(name="data")
    data_style.font = Font(name='Arial', size=10, color=Colors.DARK_TEXT)
    wb.add_named_style(data_style)

def create_dashboard_data(ws):
    """Create Dashboard worksheet with comprehensive data"""
    
    # Title
    ws['A1'] = "🏠 BEAUTY PRO DASHBOARD"
    ws['A1'].style = "header"
    ws.merge_cells('A1:J1')
    
    # Current date
    ws['A2'] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Business Overview Section
    ws['A4'] = "📊 BUSINESS OVERVIEW"
    ws['A4'].style = "subheader"
    
    # Key metrics
    metrics = [
        ("Total Products", 142, "Low Stock Items", 23),
        ("Inventory Value", "$12,847", "Expiring Soon", 5),
        ("Avg Profit Margin", "47.2%", "Stock-out Rate", "6.4%"),
        ("Inventory Turnover", "4.2x", "Expiry Risk Value", "$125")
    ]
    
    for i, (metric1, value1, metric2, value2) in enumerate(metrics, 6):
        ws[f'A{i}'] = metric1
        ws[f'B{i}'] = value1
        ws[f'D{i}'] = metric2
        ws[f'E{i}'] = value2
    
    # Quick Actions
    ws['G4'] = "📈 QUICK ACTIONS"
    ws['G4'].style = "subheader"
    
    actions = ["➕ Add Product", "🔄 Update Stock", "📊 View Reports", "📱 Mobile Entry"]
    for i, action in enumerate(actions, 6):
        ws[f'G{i}'] = action
    
    # Setup Progress
    ws['I4'] = "📋 SETUP PROGRESS"
    ws['I4'].style = "subheader"
    ws['I6'] = "Setup Complete: 85%"
    ws['I7'] = "⚠️ 2 Steps Remaining"
    ws['I8'] = "🚀 Continue Setup"
    ws['I9'] = "📖 Get Help"
    
    # Stock Status Distribution
    ws['A11'] = "📈 STOCK STATUS DISTRIBUTION"
    ws['A11'].style = "subheader"
    
    stock_status = [
        ("Healthy Stock", 89),
        ("Low Stock", 23),
        ("Out of Stock", 5),
        ("Expired", 3)
    ]
    
    for i, (status, count) in enumerate(stock_status, 13):
        ws[f'A{i}'] = status
        ws[f'B{i}'] = count
    
    # Category Performance
    ws['D11'] = "📊 CATEGORY PERFORMANCE"
    ws['D11'].style = "subheader"
    
    categories = [
        ("Lipstick", "$4,250"),
        ("Foundation", "$3,180"),
        ("Skincare", "$2,890"),
        ("Fragrance", "$1,650"),
        ("Eye Makeup", "$877"),
        ("Nail Care", "$445")
    ]
    
    for i, (category, value) in enumerate(categories, 13):
        ws[f'D{i}'] = category
        ws[f'E{i}'] = value
    
    # Alerts
    ws['G11'] = "⚠️ ALERTS & NOTIFICATIONS"
    ws['G11'].style = "subheader"
    
    alerts = [
        "🔴 5 Products Out of Stock",
        "🟡 23 Items Need Reordering",
        "⏰ 12 Items Expiring This Month",
        "📦 3 Supplier Deliveries Overdue"
    ]
    
    for i, alert in enumerate(alerts, 13):
        ws[f'G{i}'] = alert

def create_categories_data(ws):
    """Create Categories worksheet"""
    
    ws['A1'] = "📂 MY CATEGORIES"
    ws['A1'].style = "header"
    ws.merge_cells('A1:H1')
    
    # Headers
    headers = ["Category Name", "Description", "Target Margin %", "Reorder Days", "Status", "Products Count", "Last Updated", "Notes"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=header).style = "subheader"
    
    # Sample data
    categories_data = [
        ["Lipstick", "Lip colors and treatments", "45%", 14, "✅ Active", 28, "2024-01-15", "Bestselling category"],
        ["Foundation", "Base makeup products", "50%", 21, "✅ Active", 15, "2024-01-14", "High margin focus"],
        ["Skincare", "Cleansers serums moisturizers", "40%", 30, "✅ Active", 22, "2024-01-13", "Growing demand"],
        ["Fragrance", "Perfumes and body sprays", "35%", 45, "✅ Active", 12, "2024-01-12", "Luxury items"],
        ["Eye Makeup", "Eyeshadows mascara eyeliner", "42%", 18, "✅ Active", 18, "2024-01-11", "Seasonal variations"],
        ["Nail Care", "Nail polish and treatments", "38%", 25, "✅ Active", 8, "2024-01-10", "Steady performers"]
    ]
    
    for row_idx, row_data in enumerate(categories_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value).style = "data"

def create_suppliers_data(ws):
    """Create Suppliers worksheet"""
    
    ws['A1'] = "🏪 MY SUPPLIERS"
    ws['A1'].style = "header"
    ws.merge_cells('A1:J1')
    
    headers = ["Supplier Name", "Contact Person", "Email", "Phone", "Lead Time (Days)", "Payment Terms", "Categories", "Rating", "Last Order", "Notes"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=header).style = "subheader"
    
    suppliers_data = [
        ["Beauty Supply Co", "Sarah Johnson", "sarah@beautysupply.com", "(555) 123-4567", 14, "Net 30", "Lipstick Foundation", "⭐⭐⭐⭐⭐", "2024-01-10", "Reliable fast shipping"],
        ["Glamour Wholesale", "Mike Chen", "orders@glamourwholesale.com", "(555) 234-5678", 21, "Net 45", "Skincare Fragrance", "⭐⭐⭐⭐", "2024-01-08", "Good prices bulk discounts"],
        ["Premium Cosmetics", "Lisa Rodriguez", "lisa@premiumcosmetics.net", "(555) 345-6789", 18, "COD", "Eye Makeup Nail Care", "⭐⭐⭐⭐⭐", "2024-01-05", "Premium brands specialist"],
        ["Luxury Beauty Inc", "David Park", "david@luxurybeauty.com", "(555) 456-7890", 28, "Net 60", "Fragrance", "⭐⭐⭐", "2024-01-03", "High-end products only"]
    ]
    
    for row_idx, row_data in enumerate(suppliers_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value).style = "data"

def create_products_data(ws):
    """Create Products worksheet with formulas"""
    
    ws['A1'] = "💄 MY PRODUCTS"
    ws['A1'].style = "header"
    ws.merge_cells('A1:N1')
    
    headers = ["Product Name", "Brand", "Category", "SKU", "Supplier", "Cost", "Retail Price", "Margin %", "Barcode", "Expiry Date", "Min Stock", "Max Stock", "Location", "Notes"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=header).style = "subheader"
    
    products_data = [
        ["MAC Ruby Woo Lipstick", "MAC", "Lipstick", "MAC-RW-001", "Beauty Supply Co", 12.00, 24.00, "", "123456789012", "2025-12-31", 5, 25, "Shelf A1", "Bestseller"],
        ["Fenty Beauty Foundation 210", "Fenty Beauty", "Foundation", "FENTY-210", "Beauty Supply Co", 18.00, 36.00, "", "234567890123", "2025-06-30", 3, 15, "Shelf B2", "Popular shade"],
        ["The Ordinary Niacinamide Serum", "The Ordinary", "Skincare", "TO-NIAC-30", "Glamour Wholesale", 4.50, 12.00, "", "345678901234", "2025-08-15", 10, 50, "Shelf C3", "High margin"],
        ["Chanel No. 5 Eau de Parfum", "Chanel", "Fragrance", "CHANEL-5-50", "Luxury Beauty Inc", 45.00, 89.00, "", "456789012345", "2026-12-31", 2, 8, "Cabinet D1", "Luxury item"],
        ["Urban Decay Eyeshadow Palette", "Urban Decay", "Eye Makeup", "UD-NAKED-3", "Premium Cosmetics", 25.00, 54.00, "", "567890123456", "2025-10-20", 2, 10, "Shelf E4", "Seasonal favorite"],
        ["OPI Nail Polish Classic Red", "OPI", "Nail Care", "OPI-RED-15", "Premium Cosmetics", 6.00, 15.00, "", "678901234567", "2025-05-30", 5, 20, "Shelf F5", "Classic color"]
    ]
    
    for row_idx, row_data in enumerate(products_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx == 8 and row_idx > 4:  # Margin % column
                # Add formula for margin calculation
                cost_cell = get_column_letter(6) + str(row_idx)
                price_cell = get_column_letter(7) + str(row_idx)
                formula = f"=IF(AND({cost_cell}>0,{price_cell}>0),ROUND((({price_cell}-{cost_cell})/{price_cell})*100,1)&\"%\",\"\")"
                ws.cell(row=row_idx, column=col_idx, value=formula)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value).style = "data"

def create_inventory_data(ws):
    """Create Inventory worksheet with status formulas"""
    
    ws['A1'] = "📦 LIVE INVENTORY"
    ws['A1'].style = "header"
    ws.merge_cells('A1:O1')
    
    headers = ["Product Name", "Current Stock", "Min Stock", "Max Stock", "Reorder Level", "Status", "Days to Expiry", "Last Updated", "Location", "Cost", "Retail", "Total Value", "Reorder Qty", "Supplier", "Action"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=header).style = "subheader"
    
    inventory_data = [
        ["MAC Ruby Woo Lipstick", 12, 5, 25, 8, "", 365, "2024-01-15", "Shelf A1", 12.00, 24.00, "=B5*J5", "=D5-B5", "Beauty Supply Co", ""],
        ["Fenty Beauty Foundation 210", 2, 3, 15, 5, "", 181, "2024-01-14", "Shelf B2", 18.00, 36.00, "=B6*J6", "=D6-B6", "Beauty Supply Co", ""],
        ["The Ordinary Niacinamide Serum", 25, 10, 50, 15, "", 227, "2024-01-13", "Shelf C3", 4.50, 12.00, "=B7*J7", "=D7-B7", "Glamour Wholesale", ""],
        ["Chanel No. 5 Eau de Parfum", 4, 2, 8, 3, "", 730, "2024-01-12", "Cabinet D1", 45.00, 89.00, "=B8*J8", "=D8-B8", "Luxury Beauty Inc", ""],
        ["Urban Decay Eyeshadow Palette", 1, 2, 10, 3, "", 294, "2024-01-11", "Shelf E4", 25.00, 54.00, "=B9*J9", "=D9-B9", "Premium Cosmetics", ""],
        ["OPI Nail Polish Classic Red", 8, 5, 20, 8, "", 151, "2024-01-10", "Shelf F5", 6.00, 15.00, "=B10*J10", "=D10-B10", "Premium Cosmetics", ""]
    ]
    
    for row_idx, row_data in enumerate(inventory_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx == 6:  # Status column - add formula
                current_stock_cell = f"B{row_idx}"
                min_stock_cell = f"C{row_idx}"
                formula = f'=IF({current_stock_cell}=0,"🔴 Out of Stock",IF({current_stock_cell}<={min_stock_cell},"🟡 Low Stock",IF({current_stock_cell}={min_stock_cell},"🟡 At Minimum","🟢 Healthy")))'
                ws.cell(row=row_idx, column=col_idx, value=formula)
            elif col_idx == 15:  # Action column - add formula
                current_stock_cell = f"B{row_idx}"
                min_stock_cell = f"C{row_idx}"
                formula = f'=IF({current_stock_cell}=0,"URGENT ORDER",IF({current_stock_cell}<={min_stock_cell},"ORDER NOW",IF({current_stock_cell}={min_stock_cell},"Monitor","Continue")))'
                ws.cell(row=row_idx, column=col_idx, value=formula)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value).style = "data"

def create_quickadd_data(ws):
    """Create QuickAdd worksheet"""
    
    ws['A1'] = "➕ QUICK ADD INVENTORY"
    ws['A1'].style = "header"
    ws.merge_cells('A1:H1')
    
    # Quick update section
    ws['A3'] = "🚀 QUICK STOCK UPDATE"
    ws['A3'].style = "subheader"
    
    ws['A5'] = "Select Product:"
    ws['B5'] = "[Dropdown]"
    ws['D5'] = "Quantity Change:"
    ws['E5'] = 0
    ws['G5'] = "📝 Transaction Type:"
    ws['H5'] = "[Dropdown]"
    
    ws['A6'] = "Current Stock:"
    ws['B6'] = "--"
    ws['D6'] = "New Stock Level:"
    ws['E6'] = "--"
    ws['G6'] = "📅 Date:"
    ws['H6'] = "=TODAY()"
    
    ws['A7'] = "Location:"
    ws['B7'] = "--"
    ws['D7'] = "Notes:"
    ws['G7'] = "🔄 UPDATE STOCK"

def create_reorder_data(ws):
    """Create Reorder worksheet"""
    
    ws['A1'] = "🔄 REORDER DASHBOARD"
    ws['A1'].style = "header"
    ws.merge_cells('A1:J1')
    
    ws['A3'] = "📋 ITEMS NEEDING REORDER"
    ws['A3'].style = "subheader"
    
    headers = ["Product Name", "Current Stock", "Min Level", "Reorder To", "Order Qty", "Supplier", "Cost per Unit", "Total Cost", "Priority", "Action"]
    for i, header in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=header).style = "subheader"
    
    reorder_data = [
        ["Drunk Elephant Vitamin C", 0, 2, 8, 8, "Glamour Wholesale", 28.00, "=E7*G7", "🔴 URGENT", "ORDER NOW"],
        ["Fenty Beauty Foundation 210", 2, 3, 15, 13, "Beauty Supply Co", 18.00, "=E8*G8", "🔴 HIGH", "ORDER NOW"],
        ["Urban Decay Eyeshadow Palette", 1, 2, 10, 9, "Premium Cosmetics", 25.00, "=E9*G9", "🔴 HIGH", "ORDER NOW"],
        ["Charlotte Tilbury Lipstick", 3, 3, 12, 9, "Premium Cosmetics", 20.00, "=E10*G10", "🟡 MEDIUM", "Consider Order"]
    ]
    
    for row_idx, row_data in enumerate(reorder_data, 7):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value).style = "data"

def create_analytics_data(ws):
    """Create Analytics worksheet"""
    
    ws['A1'] = "📈 BUSINESS ANALYTICS"
    ws['A1'].style = "header"
    ws.merge_cells('A1:J1')
    
    # Performance Dashboard
    ws['A3'] = "📊 PERFORMANCE DASHBOARD"
    ws['A3'].style = "subheader"
    
    performance_data = [
        ["Current Month", "January 2024", "", "Previous Month", "December 2023", "", "Growth", "+12.5%"],
        ["Total Revenue", "$18,470", "", "Total Revenue", "$16,420", "", "Revenue Growth", "+$2,050"],
        ["Total Profit", "$8,730", "", "Total Profit", "$7,890", "", "Profit Growth", "+$840"],
        ["Profit Margin", "47.3%", "", "Profit Margin", "48.1%", "", "Margin Change", "-0.8%"],
        ["Units Sold", 245, "", "Units Sold", 218, "", "Volume Growth", "+27 units"]
    ]
    
    for row_idx, row_data in enumerate(performance_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value).style = "data"

def create_instructions_data(ws):
    """Create Instructions worksheet"""
    
    ws['A1'] = "📖 SETUP INSTRUCTIONS"
    ws['A1'].style = "header"
    ws.merge_cells('A1:J1')
    
    ws['A3'] = "🚀 WELCOME TO BEAUTY PRO INVENTORY SYSTEM"
    ws['A3'].style = "subheader"
    
    ws['A5'] = "🎯 QUICK START GUIDE"
    ws['A5'].style = "subheader"
    
    instructions = [
        "Step 1: CUSTOMIZE CATEGORIES (5 mins)",
        "→ Go to 📂 My Categories sheet",
        "→ Replace example categories with YOUR product types",
        "→ Set target margins and reorder days",
        "",
        "Step 2: ADD YOUR SUPPLIERS (8 mins)",
        "→ Go to 🏪 My Suppliers sheet",
        "→ Replace examples with your actual vendors",
        "→ Add complete contact information",
        "",
        "Step 3: BUILD PRODUCT CATALOG (12 mins)",
        "→ Go to 💄 My Products sheet",
        "→ Add your actual products",
        "→ Margins calculate automatically",
        "",
        "Step 4: SET CURRENT INVENTORY (5 mins)",
        "→ Go to 📦 Live Inventory sheet",
        "→ Enter current stock levels",
        "→ Status updates automatically",
        "",
        "✅ SETUP COMPLETE!",
        "Your dashboard will now show real business data"
    ]
    
    for i, instruction in enumerate(instructions, 7):
        ws[f'A{i}'] = instruction

def format_sheet(ws, sheet_name):
    """Apply formatting to each sheet"""
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders to data tables
    thin_border = Border(
        left=Side(style='thin', color=Colors.LIGHT_BORDER),
        right=Side(style='thin', color=Colors.LIGHT_BORDER),
        top=Side(style='thin', color=Colors.LIGHT_BORDER),
        bottom=Side(style='thin', color=Colors.LIGHT_BORDER)
    )
    
    # Apply borders to data ranges (this is simplified)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and cell.row > 2:  # Skip title rows
                cell.border = thin_border

if __name__ == "__main__":
    output_file = create_final_workbook()
    print(f"Final workbook created: {output_file}")